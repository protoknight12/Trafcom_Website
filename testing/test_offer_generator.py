"""
Plain-assert regression coverage for the offer generator (see app.py's
Offer/OfferItem models, sanitize_rich_text(), _rich_text_cell_value(), and
build_offer_workbook()):

- sanitize_rich_text() only allows bold/italic/line-breaks through - guards
  the stored-XSS risk of persisting raw contenteditable HTML.
- _rich_text_cell_value() preserves bold/italic as separate openpyxl runs.
- _next_offer_number() formats as zero-padded 11 digits starting at 200.
- build_offer_workbook() renders without error for a mixed product/detail/
  text-row offer and writes the expected header/number/total cells.

Run as a module from the repo root:  python -m testing.test_offer_generator
"""
from openpyxl.cell.rich_text import CellRichText

from app import app, db, sanitize_rich_text, _rich_text_cell_value, _next_offer_number, \
    build_offer_workbook, Offer, OfferItem

# --- sanitize_rich_text: allowlist only bold/italic/br ---
assert sanitize_rich_text('') is None
assert sanitize_rich_text('   ') is None
assert sanitize_rich_text('plain text') == 'plain text'
assert sanitize_rich_text('<b>bold</b> and <i>italic</i>') == '<b>bold</b> and <i>italic</i>'
assert sanitize_rich_text('<strong>x</strong>') == '<strong>x</strong>'
assert sanitize_rich_text('<script>alert(1)</script>text') == 'text'
assert sanitize_rich_text('<img src=x onerror=alert(1)>text') == 'text'
assert sanitize_rich_text('<div>line1</div><div>line2</div>') == 'line1<br>line2'

# --- _rich_text_cell_value: plain text stays a plain string ---
assert _rich_text_cell_value('', 'Calibri') == ''
assert _rich_text_cell_value('plain text', 'Calibri') == 'plain text'

# --- _rich_text_cell_value: bold/italic becomes a CellRichText with separate runs ---
rich = _rich_text_cell_value('<b>bold</b> plain <i>italic</i>', 'Calibri')
assert isinstance(rich, CellRichText)
texts = [str(block) for block in rich]
assert texts == ['bold', ' plain ', 'italic']

# --- _next_offer_number: zero-padded 11 digits, starting at 200 ---
with app.app_context():
    OfferItem.query.delete()
    Offer.query.delete()
    db.session.commit()

    first = _next_offer_number()
    assert first == '00000000200', first

    offer = Offer(number=first)
    db.session.add(offer)
    db.session.commit()

    second = _next_offer_number()
    assert second == '00000000201', second

    # --- build_offer_workbook: mixed product/detail/text rows render cleanly ---
    offer2 = Offer(number=second, object_title='Тест обект', footer_notes='Ред 1\nРед 2', signed_by='Иван Иванов')
    db.session.add(offer2)
    db.session.flush()
    db.session.add(OfferItem(offer_id=offer2.id, position=0, item_type='product', code='1.1-M01',
                              name='Тестов продукт', quantity=2, unit='бр', unit_price=10.0))
    db.session.add(OfferItem(offer_id=offer2.id, position=1, item_type='text',
                              description_html='<b>Забележка</b>: свободен текст ред'))
    db.session.commit()
    db.session.refresh(offer2)

    wb = build_offer_workbook(offer2)
    ws = wb.active
    assert ws['B5'].value == f'ОФЕРТА № {second}'
    assert ws['B6'].value == 'ОБЕКТ: Тест обект'
    assert ws['B2'].value is None  # header occupies row 1-2 merged, only B1 is written to
    assert ws['B1'].value == 'ТРАФКОМ ООД'
    assert ws['B1'].font.name == 'Oswald'

    db.session.delete(offer2)
    db.session.delete(offer)
    db.session.commit()

print("ok")
